import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import json
import os
from pathlib import Path

# Page config
st.set_page_config(
    page_title="Advanced Shift Schedule Manager",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    .stButton>button {
        width: 100%;
        border-radius: 10px;
        height: 3em;
        font-weight: bold;
    }
    .shift-card {
        background: white;
        padding: 20px;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-bottom: 20px;
    }
    div[data-testid="stMetricValue"] {
        font-size: 2em;
        font-weight: bold;
    }
    .calendar-day {
        padding: 10px;
        border: 1px solid #ddd;
        border-radius: 5px;
        min-height: 80px;
        margin: 2px;
        cursor: pointer;
        transition: all 0.3s;
    }
    .calendar-day:hover {
        box-shadow: 0 2px 8px rgba(0,0,0,0.2);
        transform: translateY(-2px);
    }
    .shift-legend {
        display: inline-block;
        padding: 5px 10px;
        margin: 5px;
        border-radius: 5px;
        color: white;
        font-weight: bold;
    }
    .info-box {
        background: rgba(255,255,255,0.9);
        padding: 15px;
        border-radius: 10px;
        margin: 10px 0;
        border-left: 4px solid #667eea;
    }
</style>
""", unsafe_allow_html=True)

# File paths for persistent storage
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
MEMBERS_FILE = DATA_DIR / "team_members.json"
SCHEDULE_FILE = DATA_DIR / "shift_schedule.json"
SETTINGS_FILE = DATA_DIR / "settings.json"
PATTERNS_FILE = DATA_DIR / "shift_patterns.json"

# Enhanced shift type definitions with SAST times
SHIFT_TYPES = {
    0: {
        'code': '',
        'name': 'Off',
        'time': 'Day Off',
        'color': '#FFFFFF',
        'text_color': '#000000'
    },
    1: {
        'code': 'D1',
        'name': 'Day Shift 1',
        'time': '7:00 AM - 4:00 PM SAST',
        'color': '#3B82F6',
        'text_color': '#FFFFFF'
    },
    2: {
        'code': 'D2',
        'name': 'Day Shift 2',
        'time': '8:00 AM - 5:00 PM SAST',
        'color': '#2563EB',
        'text_color': '#FFFFFF'
    },
    3: {
        'code': 'L',
        'name': 'Layover',
        'time': '2:00 PM - 10:00 PM SAST',
        'color': '#F59E0B',
        'text_color': '#FFFFFF'
    },
    4: {
        'code': 'N',
        'name': 'Night Shift',
        'time': '4:00 PM - 1:00 AM SAST',
        'color': '#1F2937',
        'text_color': '#FFFFFF'
    },
    5: {
        'code': 'EM',
        'name': 'Early Morning',
        'time': '3:00 AM - 11:00 AM SAST',
        'color': '#8B5CF6',
        'text_color': '#FFFFFF'
    },
    6: {
        'code': 'WD',
        'name': 'Weekend Day',
        'time': '7:00 AM - 4:00 PM SAST',
        'color': '#10B981',
        'text_color': '#FFFFFF'
    },
    7: {
        'code': 'WEM',
        'name': 'Weekend Early',
        'time': '3:00 AM - 11:00 AM SAST',
        'color': '#059669',
        'text_color': '#FFFFFF'
    },
    8: {
        'code': 'WN',
        'name': 'Weekend Night',
        'time': '4:00 PM - 1:00 AM SAST',
        'color': '#047857',
        'text_color': '#FFFFFF'
    },
    9: {
        'code': 'HD',
        'name': 'Holiday Day',
        'time': '7:00 AM - 4:00 PM SAST',
        'color': '#DC2626',
        'text_color': '#FFFFFF'
    },
    10: {
        'code': 'HEM',
        'name': 'Holiday Early',
        'time': '3:00 AM - 11:00 AM SAST',
        'color': '#B91C1C',
        'text_color': '#FFFFFF'
    },
    11: {
        'code': 'HN',
        'name': 'Holiday Night',
        'time': '4:00 PM - 1:00 AM SAST',
        'color': '#991B1B',
        'text_color': '#FFFFFF'
    },
    12: {
        'code': 'X',
        'name': 'Leave',
        'time': 'Approved Leave',
        'color': '#EC4899',
        'text_color': '#FFFFFF'
    },
    13: {
        'code': 'SL',
        'name': 'Sick Leave',
        'time': 'Sick Leave',
        'color': '#EF4444',
        'text_color': '#FFFFFF'
    },
    14: {
        'code': 'TR',
        'name': 'Training',
        'time': 'Training/Development',
        'color': '#06B6D4',
        'text_color': '#FFFFFF'
    }
}

# Data management functions
def load_team_members():
    """Load team members from JSON file"""
    if MEMBERS_FILE.exists():
        try:
            with open(MEMBERS_FILE, 'r') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
        except Exception as e:
            st.error(f"Error loading team members: {e}")
    return {}

def save_team_members(team_members):
    """Save team members to JSON file"""
    try:
        with open(MEMBERS_FILE, 'w') as f:
            json.dump(team_members, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving team members: {e}")
        return False

def load_shift_schedule():
    """Load shift schedule from JSON file"""
    if SCHEDULE_FILE.exists():
        try:
            with open(SCHEDULE_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error loading shift schedule: {e}")
    return {}

def save_shift_schedule(shift_schedule):
    """Save shift schedule to JSON file"""
    try:
        with open(SCHEDULE_FILE, 'w') as f:
            json.dump(shift_schedule, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving shift schedule: {e}")
        return False

def load_shift_patterns():
    """Load saved shift patterns"""
    if PATTERNS_FILE.exists():
        try:
            with open(PATTERNS_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error loading shift patterns: {e}")
    return {}

def save_shift_patterns(patterns):
    """Save shift patterns to JSON file"""
    try:
        with open(PATTERNS_FILE, 'w') as f:
            json.dump(patterns, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving shift patterns: {e}")
        return False

def load_settings():
    """Load app settings from JSON file"""
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error loading settings: {e}")
    return {
        'current_month': datetime.now().month,
        'current_year': datetime.now().year
    }

def save_settings(settings):
    """Save app settings to JSON file"""
    try:
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(settings, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving settings: {e}")
        return False

# Initialize session state
if 'team_members' not in st.session_state:
    st.session_state.team_members = load_team_members()

if 'shift_schedule' not in st.session_state:
    st.session_state.shift_schedule = load_shift_schedule()

if 'shift_patterns' not in st.session_state:
    st.session_state.shift_patterns = load_shift_patterns()

if 'settings' not in st.session_state:
    st.session_state.settings = load_settings()

if 'current_month' not in st.session_state:
    st.session_state.current_month = st.session_state.settings.get('current_month', datetime.now().month)

if 'current_year' not in st.session_state:
    st.session_state.current_year = st.session_state.settings.get('current_year', datetime.now().year)

# Helper functions
def get_days_in_month(year, month):
    return calendar.monthrange(year, month)[1]

def get_shift_info(shift_type):
    """Get shift information by type"""
    return SHIFT_TYPES.get(shift_type, SHIFT_TYPES[0])

def update_shift(member_name, day, shift_type):
    """Update a shift and save to file"""
    if member_name not in st.session_state.shift_schedule:
        st.session_state.shift_schedule[member_name] = [0] * 31
    
    if day < len(st.session_state.shift_schedule[member_name]):
        st.session_state.shift_schedule[member_name][day] = shift_type
        save_shift_schedule(st.session_state.shift_schedule)

def bulk_update_shifts(member_name, start_day, end_day, shift_type):
    """Update multiple days at once"""
    if member_name not in st.session_state.shift_schedule:
        st.session_state.shift_schedule[member_name] = [0] * 31
    
    for day in range(start_day, end_day + 1):
        if day < len(st.session_state.shift_schedule[member_name]):
            st.session_state.shift_schedule[member_name][day] = shift_type
    
    save_shift_schedule(st.session_state.shift_schedule)
    return True

def apply_shift_pattern(member_name, pattern, start_day=0):
    """Apply a shift pattern to a member"""
    if member_name not in st.session_state.shift_schedule:
        st.session_state.shift_schedule[member_name] = [0] * 31
    
    days_in_month = len(st.session_state.shift_schedule[member_name])
    pattern_length = len(pattern)
    
    for i in range(start_day, days_in_month):
        pattern_index = (i - start_day) % pattern_length
        st.session_state.shift_schedule[member_name][i] = pattern[pattern_index]
    
    save_shift_schedule(st.session_state.shift_schedule)
    return True

def add_team_member(team_name, member_data):
    """Add a new team member and save"""
    if team_name not in st.session_state.team_members:
        st.session_state.team_members[team_name] = []
    
    existing_names = [m['name'] for m in st.session_state.team_members[team_name]]
    if member_data['name'] in existing_names:
        return False, "Member already exists in this team"
    
    st.session_state.team_members[team_name].append(member_data)
    st.session_state.shift_schedule[member_data['name']] = [0] * 31
    
    save_team_members(st.session_state.team_members)
    save_shift_schedule(st.session_state.shift_schedule)
    
    return True, "Member added successfully"

def remove_team_member(team_name, member_name):
    """Remove a team member and save"""
    if team_name in st.session_state.team_members:
        st.session_state.team_members[team_name] = [
            m for m in st.session_state.team_members[team_name] 
            if m['name'] != member_name
        ]
        
        if member_name in st.session_state.shift_schedule:
            del st.session_state.shift_schedule[member_name]
        
        save_team_members(st.session_state.team_members)
        save_shift_schedule(st.session_state.shift_schedule)
        return True, f"Removed {member_name} from {team_name}"
    return False, "Team or member not found"

def export_to_excel():
    """Export schedule to Excel with enhanced formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule {st.session_state.current_month}-{st.session_state.current_year}"
    
    # Header styling
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    days = get_days_in_month(st.session_state.current_year, st.session_state.current_month)
    
    # Headers
    headers = ['Member', 'Team', 'Location', 'WHMCS'] + [f'Day {i}' for i in range(1, days + 1)]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 2
    for team_name, members in st.session_state.team_members.items():
        for member in members:
            ws.cell(row=row, column=1, value=member['name']).border = border
            ws.cell(row=row, column=2, value=team_name).border = border
            ws.cell(row=row, column=3, value=member['location']).border = border
            ws.cell(row=row, column=4, value=member['whmcs']).border = border
            
            schedule = st.session_state.shift_schedule.get(member['name'], [0] * days)
            for day in range(days):
                col = day + 5
                shift_type = schedule[day] if day < len(schedule) else 0
                shift_info = get_shift_info(shift_type)
                
                cell = ws.cell(row=row, column=col, value=shift_info['code'])
                cell.fill = PatternFill(start_color=shift_info['color'].replace('#', ''), 
                                       end_color=shift_info['color'].replace('#', ''), 
                                       fill_type="solid")
                cell.font = Font(bold=True, color=shift_info['text_color'].replace('#', ''))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row += 1
    
    # Add legend sheet
    legend_ws = wb.create_sheet("Legend")
    legend_ws.cell(row=1, column=1, value="Shift Code").font = Font(bold=True)
    legend_ws.cell(row=1, column=2, value="Shift Name").font = Font(bold=True)
    legend_ws.cell(row=1, column=3, value="Time").font = Font(bold=True)
    
    legend_row = 2
    for shift_type, info in SHIFT_TYPES.items():
        if shift_type > 0:  # Skip "Off"
            legend_ws.cell(row=legend_row, column=1, value=info['code'])
            legend_ws.cell(row=legend_row, column=2, value=info['name'])
            legend_ws.cell(row=legend_row, column=3, value=info['time'])
            legend_row += 1
    
    # Auto-adjust column widths
    for worksheet in wb.worksheets:
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_day_of_week(year, month, day):
    """Get day of week name"""
    date = datetime(year, month, day)
    return date.strftime('%a')

def is_weekend(year, month, day):
    """Check if day is weekend"""
    date = datetime(year, month, day)
    return date.weekday() >= 5

# Sidebar
with st.sidebar:
    st.image("https://via.placeholder.com/150x150.png?text=Shift+Manager", use_container_width=True)
    st.title("🗓️ Shift Manager")
    
    # Month/Year selector
    col1, col2 = st.columns(2)
    with col1:
        selected_month = st.selectbox(
            "Month",
            range(1, 13),
            index=st.session_state.current_month - 1,
            format_func=lambda x: calendar.month_name[x]
        )
    with col2:
        selected_year = st.selectbox(
            "Year",
            range(2024, 2031),
            index=st.session_state.current_year - 2024
        )
    
    if selected_month != st.session_state.current_month or selected_year != st.session_state.current_year:
        st.session_state.current_month = selected_month
        st.session_state.current_year = selected_year
        st.session_state.settings['current_month'] = selected_month
        st.session_state.settings['current_year'] = selected_year
        save_settings(st.session_state.settings)
    
    selected_month_name = calendar.month_name[selected_month]
    
    st.divider()
    
    # View selector
    view_type = st.radio(
        "📊 View Mode",
        ["📖 User Guide", "👥 Team Setup", "📅 Calendar View", "📊 Grid View", 
         "⚡ Bulk Assign", "🔄 Shift Patterns", "📋 Card View", "📈 Team Summary"],
        label_visibility="visible"
    )
    
    st.divider()
    
    # Export button
    if st.button("📥 Export to Excel", use_container_width=True):
        try:
            excel_data = export_to_excel()
            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_data,
                file_name=f"shift_schedule_{selected_month}_{selected_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Error exporting: {e}")
    
    # Stats
    st.divider()
    total_members = sum(len(members) for members in st.session_state.team_members.values())
    st.metric("👥 Total Members", total_members)
    st.metric("🏢 Teams", len(st.session_state.team_members))

# Main content area
if view_type == "📖 User Guide":
    st.title("📖 Advanced Shift Scheduler - User Guide")
    
    # Quick Start
    with st.expander("🚀 Quick Start Guide", expanded=True):
        st.markdown("""
        ### Getting Started in 3 Steps:
        
        1. **Add Team Members** (Team Setup)
           - Create teams and add members with their details
           - Members are automatically saved
        
        2. **Schedule Shifts** (Calendar/Grid View or Bulk Assign)
           - Use Calendar View for visual scheduling
           - Use Bulk Assign for multiple days at once
           - Use Shift Patterns for repeating schedules
        
        3. **Export & Share** (Export Button)
           - Download Excel file with full schedule
           - Includes color-coded shifts and legend
        """)
    
    # Shift Types Guide
    with st.expander("🕐 Shift Types & Times (SAST)", expanded=True):
        st.markdown("### Weekday Shifts:")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[1]["color"]}'>
                D1 - Day Shift 1
            </div>
            <p><strong>Time:</strong> 7:00 AM - 4:00 PM SAST</p>
            <p><strong>Use for:</strong> Standard morning shifts</p>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[2]["color"]}'>
                D2 - Day Shift 2
            </div>
            <p><strong>Time:</strong> 8:00 AM - 5:00 PM SAST</p>
            <p><strong>Use for:</strong> Alternative day shift</p>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[5]["color"]}'>
                EM - Early Morning
            </div>
            <p><strong>Time:</strong> 3:00 AM - 11:00 AM SAST</p>
            <p><strong>Use for:</strong> Early morning coverage</p>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[3]["color"]}'>
                L - Layover
            </div>
            <p><strong>Time:</strong> 2:00 PM - 10:00 PM SAST</p>
            <p><strong>Use for:</strong> Afternoon/evening coverage</p>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[4]["color"]}'>
                N - Night Shift
            </div>
            <p><strong>Time:</strong> 4:00 PM - 1:00 AM SAST</p>
            <p><strong>Use for:</strong> Night operations</p>
            """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("### Weekend Shifts:")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[6]["color"]}'>
                WD - Weekend Day
            </div>
            <p><strong>Time:</strong> 7:00 AM - 4:00 PM SAST</p>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[7]["color"]}'>
                WEM - Weekend Early
            </div>
            <p><strong>Time:</strong> 3:00 AM - 11:00 AM SAST</p>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[8]["color"]}'>
                WN - Weekend Night
            </div>
            <p><strong>Time:</strong> 4:00 PM - 1:00 AM SAST</p>
            """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("### Holiday Shifts:")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[9]["color"]}'>
                HD - Holiday Day
            </div>
            <p><strong>Time:</strong> 7:00 AM - 4:00 PM SAST</p>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[10]["color"]}'>
                HEM - Holiday Early
            </div>
            <p><strong>Time:</strong> 3:00 AM - 11:00 AM SAST</p>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[11]["color"]}'>
                HN - Holiday Night
            </div>
            <p><strong>Time:</strong> 4:00 PM - 1:00 AM SAST</p>
            """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("### Leave & Special:")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[12]["color"]}'>
                X - Leave
            </div>
            <p><strong>Use for:</strong> Approved time off</p>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[13]["color"]}'>
                SL - Sick Leave
            </div>
            <p><strong>Use for:</strong> Medical leave</p>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {SHIFT_TYPES[14]["color"]}'>
                TR - Training
            </div>
            <p><strong>Use for:</strong> Training days</p>
            """, unsafe_allow_html=True)
    
    # Features Guide
    with st.expander("✨ Features Overview"):
        st.markdown("""
        ### 📅 Calendar View
        - Visual month calendar with day-of-week labels
        - Click any day to see who's scheduled
        - Weekends are highlighted
        - Color-coded by shift type
        
        ### 📊 Grid View
        - See entire month schedule in table format
        - Edit individual shifts quickly
        - Filter by team
        - Export-ready format
        
        ### ⚡ Bulk Assignment
        - Assign shifts to date ranges (e.g., Monday 1st - Friday 5th)
        - Perfect for weekly schedules
        - Apply same shift to multiple days instantly
        
        ### 🔄 Shift Patterns
        - Create repeating patterns (e.g., "2 days on, 2 days off")
        - Save patterns for reuse
        - Apply to any team member
        - Examples: Night rotation, Weekend rotation
        
        ### 📥 Excel Export
        - Full schedule with color coding
        - Separate legend sheet
        - Ready for printing or sharing
        - Includes member details
        """)
    
    # Common Scenarios
    with st.expander("💡 Common Scheduling Scenarios"):
        st.markdown("""
        ### Scenario 1: Schedule a Full Week
        **Use: Bulk Assignment**
        1. Go to "Bulk Assign" view
        2. Select member
        3. Choose start day (e.g., 1) and end day (e.g., 5)
        4. Select shift type (e.g., D1)
        5. Click "Apply Bulk Assignment"
        
        ### Scenario 2: Rotating Night Shifts
        **Use: Shift Patterns**
        1. Go to "Shift Patterns" view
        2. Create pattern: e.g., "N, N, Off, Off" (2 nights, 2 off)
        3. Save pattern as "Night Rotation"
        4. Apply to member starting from Day 1
        
        ### Scenario 3: Weekend Coverage
        **Use: Calendar View**
        1. Go to "Calendar View"
        2. Navigate to weekends (highlighted)
        3. Click on Saturday/Sunday
        4. Assign WD, WEM, or WN shifts
        
        ### Scenario 4: Holiday Shifts
        **Use: Grid View or Calendar**
        1. Identify holiday dates
        2. Assign HD, HEM, or HN shifts
        3. Mark others with X (Leave) if applicable
        
        ### Scenario 5: Someone Calls in Sick
        **Use: Grid View**
        1. Go to "Grid View"
        2. Find member and day
        3. Change shift to "SL - Sick Leave"
        4. Assign replacement if needed
        """)
    
    # Tips & Best Practices
    with st.expander("🎯 Tips & Best Practices"):
        st.markdown("""
        ### Scheduling Tips:
        - ✅ **Plan ahead**: Schedule at least 2 weeks in advance
        - ✅ **Balance shifts**: Rotate night shifts fairly among team
        - ✅ **Check conflicts**: Review calendar before finalizing
        - ✅ **Mark leave early**: Update X (Leave) as soon as approved
        - ✅ **Export regularly**: Keep backup Excel files
        
        ### Efficiency Tips:
        - 🚀 Use **Bulk Assignment** for regular weekday schedules
        - 🚀 Create **Shift Patterns** for repeating rotations
        - 🚀 Use **Calendar View** for visual overview
        - 🚀 Use **Grid View** for quick individual edits
        - 🚀 Filter by team in Grid View to focus on specific groups
        
        ### Data Safety:
        - 💾 All data automatically saved to JSON files
        - 💾 Located in `data/` folder
        - 💾 Can be backed up manually
        - 💾 Persists between sessions
        """)
    
    # Troubleshooting
    with st.expander("🔧 Troubleshooting"):
        st.markdown("""
        ### Common Issues:
        
        **Q: Changes not saving?**
        - Check that you have write permissions in the `data/` folder
        - Try refreshing the page
        
        **Q: Member not showing in schedule?**
        - Make sure they're added in "Team Setup"
        - Check team filter in Grid View
        
        **Q: Calendar shows wrong month?**
        - Use month/year selector in sidebar
        - Settings are saved automatically
        
        **Q: Excel export missing shifts?**
        - Ensure shifts are saved (you should see success message)
        - Try exporting again
        
        **Q: Want to start fresh?**
        - Delete files in `data/` folder
        - Refresh the page
        """)

elif view_type == "👥 Team Setup":
    st.header("👥 Team Management")
    
    tab1, tab2 = st.tabs(["Add Members", "Remove Members"])
    
    with tab1:
        st.subheader("➕ Add New Team Member")
        
        col1, col2 = st.columns(2)
        
        with col1:
            new_team = st.text_input("Team Name", placeholder="e.g., Support Team A")
            member_name = st.text_input("Member Name", placeholder="e.g., John Doe")
        
        with col2:
            location = st.text_input("Location", placeholder="e.g., Cape Town")
            whmcs = st.text_input("WHMCS ID", placeholder="e.g., EMP001")
        
        if st.button("Add Member", type="primary", use_container_width=True):
            if new_team and member_name and location and whmcs:
                member_data = {
                    'name': member_name,
                    'location': location,
                    'whmcs': whmcs
                }
                success, message = add_team_member(new_team, member_data)
                if success:
                    st.success(f"✅ {message}")
                    st.rerun()
                else:
                    st.error(f"❌ {message}")
            else:
                st.warning("⚠️ Please fill in all fields")
        
        st.divider()
        
        # Display current teams
        st.subheader("Current Teams")
        for team_name, members in st.session_state.team_members.items():
            with st.expander(f"**{team_name}** ({len(members)} members)"):
                if members:
                    df = pd.DataFrame(members)
                    st.dataframe(df, use_container_width=True, hide_index=True)
                else:
                    st.info("No members in this team")
    
    with tab2:
        st.subheader("➖ Remove Team Member")
        
        if st.session_state.team_members:
            for team_name, members in st.session_state.team_members.items():
                if members:
                    with st.expander(f"**{team_name}**"):
                        remove_cols = st.columns([3, 1])
                        with remove_cols[0]:
                            member_to_remove = st.selectbox(
                                "Select member to remove",
                                [m['name'] for m in members],
                                key=f"remove_select_{team_name}"
                            )
                        with remove_cols[1]:
                            st.write("")
                            if st.button("Remove", key=f"remove_btn_{team_name}", type="secondary"):
                                success, message = remove_team_member(team_name, member_to_remove)
                                if success:
                                    st.success(message)
                                    st.rerun()
                                else:
                                    st.error(message)
        else:
            st.info("No teams created yet")

elif view_type == "📅 Calendar View":
    st.header(f"📅 Calendar - {selected_month_name} {selected_year}")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        # Get calendar data
        days = get_days_in_month(selected_year, selected_month)
        first_day = datetime(selected_year, selected_month, 1)
        first_weekday = first_day.weekday()
        
        # Calendar grid
        st.markdown("### Monthly Overview")
        
        # Day headers
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        header_cols = st.columns(7)
        for i, day_name in enumerate(day_names):
            with header_cols[i]:
                st.markdown(f"**{day_name}**")
        
        # Calendar days
        current_day = 1
        week_row = 0
        
        while current_day <= days:
            day_cols = st.columns(7)
            
            for col_idx in range(7):
                with day_cols[col_idx]:
                    if week_row == 0 and col_idx < first_weekday:
                        st.write("")  # Empty cell before month starts
                    elif current_day <= days:
                        is_weekend_day = is_weekend(selected_year, selected_month, current_day)
                        bg_color = "#FEE2E2" if is_weekend_day else "#F3F4F6"
                        
                        # Show day number and scheduled count
                        scheduled_today = 0
                        for member_name, schedule in st.session_state.shift_schedule.items():
                            if current_day - 1 < len(schedule) and schedule[current_day - 1] > 0:
                                scheduled_today += 1
                        
                        st.markdown(f"""
                        <div style='background-color: {bg_color}; padding: 10px; border-radius: 5px; 
                                    min-height: 60px; border: 1px solid #ddd;'>
                            <strong style='font-size: 18px;'>{current_day}</strong>
                            <br>
                            <span style='font-size: 12px; color: #666;'>
                                {scheduled_today} scheduled
                            </span>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        current_day += 1
            
            week_row += 1
        
        st.divider()
        
        # Day detail editor
        st.subheader("📝 Edit Specific Day")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            selected_day = st.number_input("Select Day", min_value=1, max_value=days, value=1)
        
        with col2:
            all_members = []
            for team_name, members in st.session_state.team_members.items():
                all_members.extend([m['name'] for m in members])
            
            if all_members:
                selected_member = st.selectbox("Select Member", all_members)
            else:
                st.warning("No members available")
                selected_member = None
        
        with col3:
            shift_options = {info['name']: shift_type for shift_type, info in SHIFT_TYPES.items()}
            selected_shift_name = st.selectbox("Shift Type", list(shift_options.keys()))
            selected_shift_type = shift_options[selected_shift_name]
        
        with col4:
            st.write("")
            st.write("")
            if selected_member and st.button("Update Shift", use_container_width=True, type="primary"):
                update_shift(selected_member, selected_day - 1, selected_shift_type)
                shift_info = get_shift_info(selected_shift_type)
                st.success(f"✅ Updated {selected_member} on Day {selected_day} to {shift_info['name']}")
                st.rerun()
        
        # Show who's scheduled for selected day
        if all_members:
            st.divider()
            st.subheader(f"Who's Working on Day {selected_day}?")
            
            day_schedule = []
            for team_name, members in st.session_state.team_members.items():
                for member in members:
                    schedule = st.session_state.shift_schedule.get(member['name'], [])
                    if selected_day - 1 < len(schedule):
                        shift_type = schedule[selected_day - 1]
                        if shift_type > 0:
                            shift_info = get_shift_info(shift_type)
                            day_schedule.append({
                                'Member': member['name'],
                                'Team': team_name,
                                'Shift': f"{shift_info['code']} - {shift_info['name']}",
                                'Time': shift_info['time']
                            })
            
            if day_schedule:
                df = pd.DataFrame(day_schedule)
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                st.info("No one scheduled for this day yet")

elif view_type == "📊 Grid View":
    st.header(f"📊 Grid Schedule - {selected_month_name} {selected_year}")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        # Stats
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Members", total_members)
        with col2:
            st.metric("Total Teams", len(st.session_state.team_members))
        with col3:
            days = get_days_in_month(selected_year, st.session_state.current_month)
            st.metric("Days in Month", days)
        with col4:
            scheduled_count = sum(
                sum(1 for shift in schedule if shift > 0)
                for schedule in st.session_state.shift_schedule.values()
            )
            st.metric("Scheduled Shifts", scheduled_count)
        
        st.divider()
        
        # Team filter
        team_filter = st.multiselect(
            "Filter by Team",
            list(st.session_state.team_members.keys()),
            default=list(st.session_state.team_members.keys())
        )
        
        # Create schedule grid
        days = get_days_in_month(selected_year, st.session_state.current_month)
        
        # Build dataframe for display
        schedule_data = []
        for team in team_filter:
            if team in st.session_state.team_members:
                for member in st.session_state.team_members[team]:
                    row = {'Member': member['name'], 'Team': team, 'Location': member['location']}
                    schedule = st.session_state.shift_schedule.get(member['name'], [0] * days)
                    for day in range(1, days + 1):
                        shift_type = schedule[day - 1] if day - 1 < len(schedule) else 0
                        shift_info = get_shift_info(shift_type)
                        row[f'Day {day}'] = shift_info['code']
                    schedule_data.append(row)
        
        if schedule_data:
            df = pd.DataFrame(schedule_data)
            
            # Style the dataframe
            def color_cells(val):
                for shift_type, info in SHIFT_TYPES.items():
                    if val == info['code']:
                        return f"background-color: {info['color']}; color: {info['text_color']}; font-weight: bold"
                return ''
            
            styled_df = df.style.applymap(
                color_cells,
                subset=[col for col in df.columns if col.startswith('Day')]
            )
            st.dataframe(styled_df, use_container_width=True, height=600)
            
            # Quick shift editor
            st.divider()
            st.subheader("✏️ Quick Edit")
            
            edit_col1, edit_col2, edit_col3, edit_col4 = st.columns(4)
            
            with edit_col1:
                all_members = []
                for team in team_filter:
                    if team in st.session_state.team_members:
                        all_members.extend([m['name'] for m in st.session_state.team_members[team]])
                
                if all_members:
                    selected_member = st.selectbox("Select Member", all_members)
                else:
                    st.warning("No members in selected teams")
                    selected_member = None
            
            with edit_col2:
                selected_day = st.number_input("Day", min_value=1, max_value=days, value=1)
            
            with edit_col3:
                shift_options = {info['name']: shift_type for shift_type, info in SHIFT_TYPES.items()}
                selected_shift_name = st.selectbox("Shift Type", list(shift_options.keys()))
                selected_shift_type = shift_options[selected_shift_name]
            
            with edit_col4:
                st.write("")
                st.write("")
                if selected_member and st.button("Update Shift", use_container_width=True, type="primary"):
                    update_shift(selected_member, selected_day - 1, selected_shift_type)
                    shift_info = get_shift_info(selected_shift_type)
                    st.success(f"✅ Updated {selected_member}'s shift for day {selected_day}")
                    st.rerun()
        else:
            st.info("No team members selected. Please select teams from the filter above.")
        
        # Legend
        st.divider()
        st.subheader("📋 Shift Legend")
        
        legend_cols = st.columns(5)
        col_idx = 0
        
        for shift_type, info in SHIFT_TYPES.items():
            if shift_type > 0:  # Skip "Off"
                with legend_cols[col_idx % 5]:
                    st.markdown(f"""
                    <div class='shift-legend' style='background-color: {info["color"]}; color: {info["text_color"]}'>
                        {info['code']}
                    </div>
                    <small><strong>{info['name']}</strong><br>{info['time']}</small>
                    """, unsafe_allow_html=True)
                col_idx += 1

elif view_type == "⚡ Bulk Assign":
    st.header("⚡ Bulk Shift Assignment")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        st.markdown("""
        <div class='info-box'>
            <strong>💡 Tip:</strong> Use bulk assignment to quickly schedule multiple consecutive days 
            for a team member. Perfect for weekly schedules or extended assignments.
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Member selection
            all_members = []
            for team_name, members in st.session_state.team_members.items():
                all_members.extend([f"{m['name']} ({team_name})" for m in members])
            
            selected_member_display = st.selectbox("Select Member", all_members)
            selected_member = selected_member_display.split(' (')[0]  # Extract name
            
            # Date range
            days = get_days_in_month(selected_year, st.session_state.current_month)
            
            date_col1, date_col2 = st.columns(2)
            with date_col1:
                start_day = st.number_input("Start Day", min_value=1, max_value=days, value=1)
            with date_col2:
                end_day = st.number_input("End Day", min_value=start_day, max_value=days, value=min(start_day + 4, days))
            
            st.info(f"📅 Will assign days {start_day} through {end_day} ({end_day - start_day + 1} days)")
        
        with col2:
            # Shift selection
            shift_options = {info['name']: shift_type for shift_type, info in SHIFT_TYPES.items()}
            selected_shift_name = st.selectbox("Shift Type", list(shift_options.keys()))
            selected_shift_type = shift_options[selected_shift_name]
            
            # Show shift details
            shift_info = get_shift_info(selected_shift_type)
            st.markdown(f"""
            <div class='shift-legend' style='background-color: {shift_info["color"]}; color: {shift_info["text_color"]}'>
                {shift_info['code']} - {shift_info['name']}
            </div>
            <p><strong>Time:</strong> {shift_info['time']}</p>
            """, unsafe_allow_html=True)
        
        st.divider()
        
        # Preview
        st.subheader("📋 Preview")
        preview_data = []
        for day in range(start_day, end_day + 1):
            day_of_week = get_day_of_week(selected_year, selected_month, day)
            preview_data.append({
                'Day': day,
                'Day of Week': day_of_week,
                'Shift Code': shift_info['code'],
                'Shift Name': shift_info['name'],
                'Time': shift_info['time']
            })
        
        preview_df = pd.DataFrame(preview_data)
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
        
        # Apply button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("✅ Apply Bulk Assignment", use_container_width=True, type="primary"):
                success = bulk_update_shifts(selected_member, start_day - 1, end_day - 1, selected_shift_type)
                if success:
                    st.success(f"✅ Successfully assigned {shift_info['name']} to {selected_member} for days {start_day}-{end_day}")
                    st.balloons()
                    st.rerun()
                else:
                    st.error("❌ Failed to apply bulk assignment")

elif view_type == "🔄 Shift Patterns":
    st.header("🔄 Shift Pattern Manager")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        st.markdown("""
        <div class='info-box'>
            <strong>💡 What are Shift Patterns?</strong><br>
            Patterns are repeating sequences of shifts that automatically cycle through the month.
            For example: "2 days on, 2 days off" or "Day, Day, Night, Night, Off, Off"
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["Create Pattern", "Apply Pattern"])
        
        with tab1:
            st.subheader("➕ Create New Pattern")
            
            pattern_name = st.text_input("Pattern Name", placeholder="e.g., 2-2 Rotation, Night Cycle")
            
            st.markdown("**Build Your Pattern:**")
            st.caption("Add shifts in the order they should repeat")
            
            # Pattern builder
            if 'temp_pattern' not in st.session_state:
                st.session_state.temp_pattern = []
            
            col1, col2 = st.columns([3, 1])
            
            with col1:
                shift_options = {info['name']: shift_type for shift_type, info in SHIFT_TYPES.items()}
                shift_to_add = st.selectbox("Select Shift to Add", list(shift_options.keys()))
            
            with col2:
                st.write("")
                st.write("")
                if st.button("➕ Add", use_container_width=True):
                    st.session_state.temp_pattern.append(shift_options[shift_to_add])
                    st.rerun()
            
            # Display current pattern
            if st.session_state.temp_pattern:
                st.markdown("**Current Pattern:**")
                pattern_display = []
                for idx, shift_type in enumerate(st.session_state.temp_pattern):
                    shift_info = get_shift_info(shift_type)
                    pattern_display.append(f"Day {idx + 1}: {shift_info['code']} - {shift_info['name']}")
                
                for item in pattern_display:
                    st.text(item)
                
                st.info(f"✅ Pattern length: {len(st.session_state.temp_pattern)} days")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("💾 Save Pattern", use_container_width=True, type="primary"):
                        if pattern_name:
                            st.session_state.shift_patterns[pattern_name] = st.session_state.temp_pattern.copy()
                            save_shift_patterns(st.session_state.shift_patterns)
                            st.success(f"✅ Pattern '{pattern_name}' saved!")
                            st.session_state.temp_pattern = []
                            st.rerun()
                        else:
                            st.error("Please enter a pattern name")
                
                with col2:
                    if st.button("🗑️ Clear Pattern", use_container_width=True):
                        st.session_state.temp_pattern = []
                        st.rerun()
            else:
                st.info("👆 Add shifts to build your pattern")
        
        with tab2:
            st.subheader("📤 Apply Saved Pattern")
            
            if st.session_state.shift_patterns:
                col1, col2 = st.columns(2)
                
                with col1:
                    # Member selection
                    all_members = []
                    for team_name, members in st.session_state.team_members.items():
                        all_members.extend([f"{m['name']} ({team_name})" for m in members])
                    
                    selected_member_display = st.selectbox("Select Member", all_members, key="apply_member")
                    selected_member = selected_member_display.split(' (')[0]
                    
                    # Pattern selection
                    pattern_name = st.selectbox("Select Pattern", list(st.session_state.shift_patterns.keys()))
                    
                    # Start day
                    days = get_days_in_month(selected_year, st.session_state.current_month)
                    start_day = st.number_input("Start from Day", min_value=1, max_value=days, value=1, key="pattern_start")
                
                with col2:
                    # Preview pattern
                    st.markdown("**Pattern Preview:**")
                    selected_pattern = st.session_state.shift_patterns[pattern_name]
                    
                    preview_days = min(14, days - start_day + 1)  # Show first 14 days or remaining days
                    
                    for i in range(preview_days):
                        pattern_idx = i % len(selected_pattern)
                        shift_type = selected_pattern[pattern_idx]
                        shift_info = get_shift_info(shift_type)
                        day_num = start_day + i
                        
                        st.markdown(f"""
                        <div class='shift-legend' style='background-color: {shift_info["color"]}; 
                                    color: {shift_info["text_color"]}; margin: 2px 0;'>
                            Day {day_num}: {shift_info['code']} - {shift_info['name']}
                        </div>
                        """, unsafe_allow_html=True)
                    
                    if preview_days < days - start_day + 1:
                        st.caption(f"... pattern continues for {days - start_day + 1} total days")
                
                st.divider()
                
                # Apply button
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    if st.button("✅ Apply Pattern", use_container_width=True, type="primary"):
                        success = apply_shift_pattern(selected_member, selected_pattern, start_day - 1)
                        if success:
                            st.success(f"✅ Applied pattern '{pattern_name}' to {selected_member} starting from day {start_day}")
                            st.balloons()
                            st.rerun()
                        else:
                            st.error("❌ Failed to apply pattern")
                
                st.divider()
                
                # Manage saved patterns
                st.subheader("📚 Saved Patterns")
                for pname, pattern in st.session_state.shift_patterns.items():
                    with st.expander(f"**{pname}** ({len(pattern)} days cycle)"):
                        pattern_display = []
                        for idx, shift_type in enumerate(pattern):
                            shift_info = get_shift_info(shift_type)
                            pattern_display.append({
                                'Position': idx + 1,
                                'Code': shift_info['code'],
                                'Shift': shift_info['name'],
                                'Time': shift_info['time']
                            })
                        
                        df = pd.DataFrame(pattern_display)
                        st.dataframe(df, use_container_width=True, hide_index=True)
                        
                        if st.button(f"🗑️ Delete '{pname}'", key=f"del_{pname}"):
                            del st.session_state.shift_patterns[pname]
                            save_shift_patterns(st.session_state.shift_patterns)
                            st.success(f"Deleted pattern '{pname}'")
                            st.rerun()
            else:
                st.info("No saved patterns yet. Create one in the 'Create Pattern' tab!")

elif view_type == "📋 Card View":
    st.header(f"📋 Team Overview - {selected_month_name} {selected_year}")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        for team_name, members in st.session_state.team_members.items():
            st.subheader(f"👥 {team_name}")
            
            member_cols = st.columns(3)
            
            for idx, member in enumerate(members):
                with member_cols[idx % 3]:
                    # Count scheduled shifts
                    schedule = st.session_state.shift_schedule.get(member['name'], [])
                    scheduled_count = sum(1 for shift in schedule if shift > 0)
                    
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                padding: 15px; border-radius: 10px; margin-bottom: 10px;'>
                        <h4 style='color: white; margin: 0;'>👤 {member['name']}</h4>
                        <p style='color: white; margin: 5px 0; font-size: 14px;'>
                            <strong>Team:</strong> {team_name}</p>
                        <p style='color: white; margin: 5px 0; font-size: 14px;'>
                            <strong>Location:</strong> {member['location']}</p>
                        <p style='color: white; margin: 5px 0; font-size: 14px;'>
                            <strong>WHMCS:</strong> {member['whmcs']}</p>
                        <p style='color: white; margin: 5px 0; font-size: 14px;'>
                            <strong>Scheduled:</strong> {scheduled_count} days this month</p>
                    </div>
                    """, unsafe_allow_html=True)
            
            st.divider()

elif view_type == "📈 Team Summary":
    st.header("📈 Team Statistics & Analytics")
    
    if total_members == 0:
        st.warning("No team members added yet. Go to 'Team Setup' to add members.")
    else:
        # Overall stats
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Teams", len(st.session_state.team_members))
        with col2:
            st.metric("Total Members", total_members)
        with col3:
            st.metric("Active Month", f"{selected_month_name} {selected_year}")
        with col4:
            scheduled_count = sum(
                sum(1 for shift in schedule if shift > 0)
                for schedule in st.session_state.shift_schedule.values()
            )
            st.metric("Total Scheduled Shifts", scheduled_count)
        
        st.divider()
        
        # Shift type breakdown
        st.subheader("📊 Shift Type Distribution")
        
        shift_counts = {}
        for shift_type, info in SHIFT_TYPES.items():
            if shift_type > 0:
                shift_counts[info['name']] = 0
        
        for schedule in st.session_state.shift_schedule.values():
            for shift in schedule:
                if shift > 0:
                    info = get_shift_info(shift)
                    shift_counts[info['name']] = shift_counts.get(info['name'], 0) + 1
        
        if any(count > 0 for count in shift_counts.values()):
            chart_data = pd.DataFrame({
                'Shift Type': list(shift_counts.keys()),
                'Count': list(shift_counts.values())
            })
            st.bar_chart(chart_data.set_index('Shift Type'))
        else:
            st.info("No shifts scheduled yet")
        
        st.divider()
        
        # Team breakdown
        st.subheader("👥 Team Breakdown")
        
        for team, members in st.session_state.team_members.items():
            with st.expander(f"**{team}** ({len(members)} members)", expanded=True):
                if len(members) > 0:
                    # Member details
                    team_df = pd.DataFrame(members)
                    st.dataframe(team_df, use_container_width=True, hide_index=True)
                    
                    # Shift statistics for this team
                    st.markdown("**Shift Statistics:**")
                    team_shifts = {}
                    for member in members:
                        schedule = st.session_state.shift_schedule.get(member['name'], [])
                        total_shifts = sum(1 for shift in schedule if shift > 0)
                        team_shifts[member['name']] = total_shifts
                    
                    if team_shifts:
                        shift_df = pd.DataFrame(
                            list(team_shifts.items()),
                            columns=['Member', 'Scheduled Days']
                        )
                        st.bar_chart(shift_df.set_index('Member'))
                else:
                    st.info("No members in this team")

# Footer
st.divider()
st.markdown(f"""
<div style='text-align: center; color: #fff; padding: 20px;'>
    <p><strong>Advanced Shift Schedule Manager v3.0</strong> | Enhanced with SAST Time Zones</p>
    <p>💾 Auto-saved | 📅 {selected_month_name} {selected_year} | 👥 {total_members} Members</p>
    <p>📊 Features: Bulk Assignment • Shift Patterns • Calendar View • Excel Export</p>
</div>
""", unsafe_allow_html=True)
